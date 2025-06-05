from flask import Flask, request, send_file
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

app = Flask(__name__)

@app.route("/generate-xlsx", methods=["POST"])
def generate_xlsx():
    """
    Генерирует XLSX файл с кастомным форматированием на основе JSON данных.
    """
    try:
        data = request.json.get("data", [])
        if not data:
            return "JSON 'data' field is missing or empty.", 400
            
        df = pd.DataFrame(data)

        # Убедимся, что все названия столбцов строковые для дальнейшей обработки
        df.columns = [str(col) for col in df.columns]

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            ws = writer.sheets['Sheet1']

            # --- 1. ОПРЕДЕЛЕНИЕ СТИЛЕЙ ---

            # Шрифты
            red_header_font = Font(name='Calibri', size=11, bold=True, color='FF0000')
            black_header_font = Font(name='Calibri', size=11, bold=True, color='000000')
            gothic_9_font = Font(name='Century Gothic', size=9)
            calibri_10_font = Font(name='Calibri', size=10)
            calibri_11_font = Font(name='Calibri', size=11)
            arial_9_font = Font(name='Arial', size=9)
            times_new_roman_9_font = Font(name='Times New Roman', size=9)

            # Заливка
            light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            bright_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

            # Границы
            dark_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Выравнивание
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')


            # --- 2. ФОРМАТИРОВАНИЕ ЗАГОЛОВКОВ ---
            
            headers_red = ['наименование', 'артикул', 'баркод', 'кол-во', 'цена поставки', 'розничная цена']
            
            # Проходим по ячейкам заголовка (первая строка)
            for cell in ws[1]:
                header_text = str(cell.value).lower()
                
                # Применяем красный или черный шрифт
                if header_text in headers_red:
                    cell.font = red_header_font
                else:
                    cell.font = black_header_font
                
                # Все заголовки - большими буквами
                cell.value = str(cell.value).upper()
                cell.alignment = center_align


            # --- 3. ФОРМАТИРОВАНИЕ ДАННЫХ ---
            
            # Получаем список заголовков уже после их преобразования в верхний регистр
            headers = [str(cell.value).lower() for cell in ws[1]]

            # Проходим по всем строкам с данными
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
                for cell in row:
                    # Определяем название столбца для текущей ячейки
                    col_header = headers[cell.col_idx - 1]

                    # Применяем стили в зависимости от столбца
                    if col_header == 'наименование':
                        cell.font = gothic_9_font
                        cell.border = dark_border
                        cell.alignment = left_align
                    
                    elif col_header == 'артикул':
                        cell.font = calibri_10_font
                        # нет темных границ
                        cell.alignment = left_align

                    elif col_header == 'баркод':
                        cell.font = calibri_11_font
                        cell.fill = light_green_fill
                        cell.alignment = center_align

                    elif col_header == 'кол-во':
                        cell.font = gothic_9_font
                        cell.fill = bright_green_fill
                        cell.border = dark_border
                        cell.alignment = center_align

                    elif col_header == 'v_размер':
                        cell.font = calibri_11_font
                        cell.alignment = left_align

                    elif col_header == 'цена поставки':
                        cell.font = arial_9_font
                        cell.alignment = left_align
                        cell.number_format = "0.00"

                    elif col_header == 'розничная цена':
                        cell.font = calibri_11_font
                        cell.alignment = left_align
                        cell.number_format = "0.00"

                    elif col_header in ['категория', 'бренд', 'единицы измерения', 'поставщик']:
                        cell.font = calibri_11_font
                        cell.alignment = left_align

                    elif col_header == 'v_цвет':
                        cell.font = gothic_9_font
                        cell.border = dark_border
                        cell.alignment = left_align

                    elif col_header == 'mark_code':
                        cell.font = calibri_11_font
                        cell.alignment = left_align
                        
                    elif col_header == 'цена usd':
                        cell.font = times_new_roman_9_font
                        cell.border = dark_border
                        cell.alignment = left_align
                        cell.number_format = "0.00"

                    else:
                        # Стиль по умолчанию для столбцов, которые не были описаны
                        cell.font = calibri_11_font
                        cell.alignment = left_align
        
        output.seek(0)
        return send_file(
            output,
            download_name="formatted_table.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return str(e), 500


if __name__ == "__main__":
    # Для запуска этого сервера нужны Flask и pandas:
    # pip install Flask pandas openpyxl
    app.run(host="0.0.0.0", port=10000)
